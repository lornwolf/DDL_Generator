#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""DDL Generator - Excel仕様書からMySQL DDLを生成するCLIツール (source.vba Python移植版)。

依存: openpyxl  (pip install openpyxl)

使い方:
    python ddl_generator.py -i <入力フォルダ> -o <出力フォルダ> [-y]
    python ddl_generator.py -n <表ID> [<表ID> ...] -o <出力フォルダ> [-y]
"""

import argparse
import hashlib
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import load_workbook


DATETIME_PATTERN = re.compile(r"^datetime\((.*)\)$", re.IGNORECASE)

# Box Drive 同期フォルダ（テーブル設計書の置き場所）
BOX_FOLDER = Path(
    r"D:\Users\XRMTUser13\Box\C_クロッシングプロジェクト\05.基幹本体構築"
    r"\02_成果集\020_プロト開発\01_作成物\120_新テーブル設計書"
)
BOX_FILE_PREFIX = "テーブル設計書_"


def convert_to_mysql_type(field_type: str) -> str:
    t = field_type.strip().lower()
    mapping = {
        "varchar": "VARCHAR", "nvarchar": "VARCHAR", "char": "VARCHAR", "nchar": "VARCHAR",
        "text": "TEXT", "ntext": "TEXT",
        "int": "INT", "integer": "INT",
        "bigint": "BIGINT",
        "smallint": "SMALLINT",
        "tinyint": "TINYINT",
        "decimal": "DECIMAL", "numeric": "DECIMAL",
        "float": "FLOAT", "real": "FLOAT",
        "double": "DOUBLE",
        "date": "DATE",
        "datetime": "DATETIME", "timestamp": "DATETIME",
        "time": "TIME",
        "blob": "BLOB", "binary": "BLOB", "varbinary": "BLOB",
        "longblob": "LONGBLOB",
        "bit": "BIT",
        "boolean": "TINYINT", "bool": "TINYINT",
    }
    return mapping.get(t, "")


def build_mysql_type(sql_type: str, field_len: str, field_dec: str) -> str:
    if sql_type in ("VARCHAR", "CHAR"):
        return f"{sql_type}({field_len})" if field_len else f"{sql_type}(255)"
    if sql_type == "DECIMAL":
        if field_len and field_dec:
            return f"{sql_type}({field_len},{field_dec})"
        if field_len:
            return f"{sql_type}({field_len},0)"
        return f"{sql_type}(10,0)"
    if sql_type == "BIT":
        return f"{sql_type}({field_len})" if field_len else f"{sql_type}(1)"
    if sql_type == "DATETIME":
        if field_len and field_len != "0":
            return f"{sql_type}({field_len})"
        return sql_type
    return sql_type


def get_file_sha256(file_path: Path) -> str:
    h = hashlib.sha256()
    with file_path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def write_utf8_no_bom(file_path: Path, content: str) -> None:
    normalized = content.replace("\r\n", "\n").replace("\n", "\r\n")
    file_path.write_bytes(normalized.encode("utf-8"))


def cell_str(ws, row: int, col: int) -> str:
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def is_numeric(s: str) -> bool:
    if not s:
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False


def prepare_output_folder(output_folder: Path) -> bool:
    """出力フォルダを準備する。存在しなければ作成、存在すれば中身を空にする。"""
    if output_folder.exists():
        if not output_folder.is_dir():
            print(
                f"エラー：出力用パスはディレクトリではありません: {output_folder}",
                file=sys.stderr,
            )
            return False
        for item in output_folder.iterdir():
            try:
                if item.is_dir() and not item.is_symlink():
                    shutil.rmtree(item)
                else:
                    item.unlink()
            except OSError as e:
                print(
                    f"エラー：出力フォルダ内のアイテム削除に失敗しました: {item} ({e})",
                    file=sys.stderr,
                )
                return False
    else:
        try:
            output_folder.mkdir(parents=True, exist_ok=True)
        except OSError as e:
            print(
                f"エラー：出力フォルダの作成に失敗しました: {output_folder} ({e})",
                file=sys.stderr,
            )
            return False
    return True


def resolve_box_files(table_ids: List[str], warnings: List[str],
                      errors: List[str]) -> List[Path]:
    """表IDリストから Box フォルダ内の対応Excelファイルを解決する。
    見つからない表IDは errors に記録し、見つかった分だけ返す。"""
    if not BOX_FOLDER.is_dir():
        errors.append(f"Boxフォルダが見つかりません: {BOX_FOLDER}")
        return []

    candidates = [
        p for p in BOX_FOLDER.iterdir()
        if p.is_file()
        and p.suffix.lower() == ".xlsx"
        and not p.name.startswith("~$")
        and p.name.startswith(BOX_FILE_PREFIX)
    ]

    resolved: List[Path] = []
    seen_paths = set()
    for tid in table_ids:
        prefix = f"{BOX_FILE_PREFIX}{tid}"
        matches = sorted(p for p in candidates if p.name.startswith(prefix))
        if not matches:
            errors.append(
                f"表ID '{tid}' に対応するExcelファイルがBoxフォルダに見つかりません: {prefix}*.xlsx"
            )
            continue
        if len(matches) > 1:
            warnings.append(
                f"表ID '{tid}' に複数のExcelファイルが見つかりました。最初のファイルを使用します: "
                + ", ".join(m.name for m in matches)
            )
        chosen = matches[0]
        if chosen in seen_paths:
            continue
        seen_paths.add(chosen)
        resolved.append(chosen)
    return resolved


def process_excel(file_path: Path, single_ddl_folder: Path,
                  warnings: List[str], errors: List[str]) -> str:
    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
    except Exception as e:
        errors.append(f"ファイル {file_path.name} の読み込みに失敗しました: {e}")
        return ""

    try:
        if len(wb.worksheets) < 2:
            warnings.append(f"ファイル {file_path.name} はSheetが1つしかないため、スキップしました！")
            return ""

        ws = wb.worksheets[1]

        table_id = cell_str(ws, 2, 4)
        table_name = cell_str(ws, 2, 5)

        if not table_id or not table_name:
            errors.append(f"ファイル {file_path.name} のテーブルIDまたはテーブル名が空です！")
            return ""

        source_hash = get_file_sha256(file_path)
        generated_at = datetime.now().strftime("%Y-%m-%d")

        has_ai_col = cell_str(ws, 5, 11).upper() == "AUTO_INCREMENT"

        primary_keys: List[str] = []
        column_defs: List[str] = []

        i = 6
        while True:
            field_name = cell_str(ws, i, 4)
            field_id = cell_str(ws, i, 5)
            if not field_name and not field_id:
                break

            pk_mark = cell_str(ws, i, 2)
            if pk_mark and pk_mark != "PK" and not pk_mark.startswith("P"):
                warnings.append(
                    f"ファイル {file_path.name} の{i}行目のB列に無効な値 '{pk_mark}' が含まれています。無視しました！"
                )
                pk_mark = ""
            is_pk = pk_mark == "PK" or (bool(pk_mark) and pk_mark.startswith("P"))

            null_mark = cell_str(ws, i, 3)
            if null_mark and null_mark != "Y":
                warnings.append(
                    f"ファイル {file_path.name} の{i}行目のC列に無効な値 '{null_mark}' が含まれています。無視しました！"
                )
                null_mark = ""
            is_nullable = (null_mark == "Y")

            is_auto_increment = False
            if has_ai_col:
                ai_mark = cell_str(ws, i, 11)
                if ai_mark and ai_mark != "〇":
                    warnings.append(
                        f"ファイル {file_path.name} の{i}行目のK列に無効な値 '{ai_mark}' が含まれています。無視しました！"
                    )
                    ai_mark = ""
                is_auto_increment = (ai_mark == "〇")

            if not field_name:
                errors.append(
                    f"ファイル {file_path.name} の{i}行目のフィールド名が空です。処理中止！"
                )
                return ""
            if not field_id:
                errors.append(
                    f"ファイル {file_path.name} の{i}行目のフィールドIDが空です。処理中止！"
                )
                return ""
            if " " in field_id:
                errors.append(
                    f"ファイル {file_path.name} の{i}行目のフィールドIDにスペースが含まれています。処理中止！"
                )
                return ""

            field_type = cell_str(ws, i, 6)
            if not field_type:
                errors.append(
                    f"ファイル {file_path.name} の{i}行目のフィールドタイプが空です。処理中止！"
                )
                return ""

            datetime_precision_from_type = ""
            m = DATETIME_PATTERN.match(field_type)
            if m:
                datetime_precision_from_type = m.group(1).strip() or "0"
                field_type = "DATETIME"

            sql_type = convert_to_mysql_type(field_type)
            if not sql_type:
                errors.append(
                    f"ファイル {file_path.name} の{i}行目のフィールドタイプ '{field_type}' はMySQLでサポートされていません。処理中止！"
                )
                return ""

            field_len = cell_str(ws, i, 7)
            if sql_type == "DATETIME":
                if datetime_precision_from_type != "":
                    field_len = datetime_precision_from_type
                if not is_numeric(field_len):
                    field_len = "0"
                else:
                    p = int(float(field_len))
                    field_len = str(p) if 0 <= p <= 6 else "0"
            else:
                if field_len and not is_numeric(field_len):
                    errors.append(
                        f"ファイル {file_path.name} の{i}行目のフィールド長が有効数字ではありません。処理中止！"
                    )
                    return ""

            field_dec = cell_str(ws, i, 8)
            if field_dec and not is_numeric(field_dec):
                errors.append(
                    f"ファイル {file_path.name} の{i}行目の小数点以下桁数が有効数字ではありません。処理中止！"
                )
                return ""

            mysql_type = build_mysql_type(sql_type, field_len, field_dec)

            if is_auto_increment:
                column_type_str = f"{mysql_type} UNSIGNED"
                column_constraints = "NOT NULL AUTO_INCREMENT"
            else:
                column_type_str = mysql_type
                column_constraints = "NULL" if is_nullable else "NOT NULL"

            col_def = f"    {field_id} {column_type_str} {column_constraints}"
            if field_name:
                col_def += f" COMMENT '{field_name.replace(chr(39), chr(39) * 2)}'"

            column_defs.append(col_def)

            if is_pk:
                primary_keys.append(field_id)

            i += 1

        if not column_defs:
            warnings.append(f"ファイル {file_path.name} は有効なフィールドがありません。スキップしました！")
            return ""

        body_lines = [",\n".join(column_defs)]
        if primary_keys:
            body_lines.append(",\n    PRIMARY KEY (" + ", ".join(primary_keys) + ")")
        body = "".join(body_lines) + "\n"

        table_comment = table_name.replace("'", "''")
        footer = f") ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='{table_comment}';"

        common_header = (
            f"-- Table: {table_id} ({table_name})\n"
            f"DROP TABLE IF EXISTS {table_id};\n"
            f"CREATE TABLE {table_id} (\n"
        )
        ddl_block = common_header + body + footer + "\n\n"

        single_block = (
            f"-- @source-hash: {source_hash}\n"
            f"-- @generated-at: {generated_at}\n"
            + common_header + body + footer + "\n"
        )

        out_path = single_ddl_folder / f"{table_id}_{table_name}.sql"
        write_utf8_no_bom(out_path, single_block)

        return ddl_block
    finally:
        wb.close()


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Excel仕様書からMySQL DDLを生成するツール (source.vba Python移植版)"
    )
    src_group = parser.add_mutually_exclusive_group(required=True)
    src_group.add_argument(
        "--input", "-i",
        help="入力フォルダ (.xlsxファイルが置かれているフォルダ)"
    )
    src_group.add_argument(
        "--names", "-n", nargs="+", metavar="TABLE_ID",
        help="Boxフォルダから取得する表IDの一覧（複数指定可）"
    )
    parser.add_argument("--output", "-o", required=True, help="出力フォルダ")
    parser.add_argument("--yes", "-y", action="store_true", help="確認プロンプトをスキップ")
    args = parser.parse_args()

    output_folder = Path(args.output)
    if output_folder.exists() and not output_folder.is_dir():
        print(
            f"エラー：出力用パスはディレクトリではありません: {output_folder}",
            file=sys.stderr,
        )
        return 1

    warnings: List[str] = []
    errors: List[str] = []

    if args.names:
        xlsx_files = resolve_box_files(args.names, warnings, errors)
        for err in errors:
            print(err, file=sys.stderr)
        if not xlsx_files:
            print("エラー：処理対象のExcelファイルがありません！", file=sys.stderr)
            return 1
        print(f"Boxフォルダから {len(xlsx_files)} 件のExcelファイルを検出しました。")
        for p in xlsx_files:
            print(f"  - {p.name}")
    else:
        input_folder = Path(args.input)
        if not input_folder.is_dir():
            print(f"エラー：入力用フォルダのパスが無効または存在しません！: {input_folder}", file=sys.stderr)
            return 1
        xlsx_files = sorted(
            p for p in input_folder.iterdir()
            if p.is_file() and p.suffix.lower() == ".xlsx" and not p.name.startswith("~$")
        )
        if not xlsx_files:
            print("エラー：入力用フォルダにExcelファイル(.xlsx)がありません！", file=sys.stderr)
            return 1

    if not args.yes:
        if output_folder.exists() and any(output_folder.iterdir()):
            print(f"注意：出力フォルダ '{output_folder}' の既存ファイルはすべて削除されます。")
        ans = input("DDL生成を開始しますか？ [y/N]: ").strip().lower()
        if ans not in ("y", "yes"):
            print("中止しました。")
            return 0

    if not prepare_output_folder(output_folder):
        return 1

    single_ddl_folder = output_folder / "CREATE文（テーブル単位）"
    single_ddl_folder.mkdir(parents=True, exist_ok=True)

    ddl_content = ""
    file_count = 0

    for f in xlsx_files:
        file_count += 1
        block = process_excel(f, single_ddl_folder, warnings, errors)
        if block:
            ddl_content += block

    if ddl_content:
        write_utf8_no_bom(output_folder / "CREATE文.sql", ddl_content)
    if warnings:
        write_utf8_no_bom(output_folder / "WARN.log", "\n".join(warnings) + "\n")
    if errors:
        write_utf8_no_bom(output_folder / "ERROR.log", "\n".join(errors) + "\n")

    print()
    print("処理完了！")
    print(f"合計 {file_count} 件のExcelファイルを処理しました。")
    if ddl_content:
        print("DDL出力先: CREATE文.sql")
    if warnings:
        print("警告内容: WARN.log")
    if errors:
        print("エラー内容: ERROR.log")
    return 0


if __name__ == "__main__":
    sys.exit(main())
